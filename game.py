import cv2
import tkinter as tk
from tkinter import messagebox
import random
import time
import sys
import mediapipe as mp
import math
import threading
import winsound
import threading
import os
import dlib
from imutils import face_utils
from scipy.spatial import distance
import openpyxl as xl


landmark_line_ids = [(0, 1), (1, 5), (5, 9), (9, 13), (13, 17), (17, 0),(1, 2), (2, 3), (3, 4), (5, 6), (6, 7), (7, 8),(9, 10), (10, 11), (11, 12), (13, 14), (14, 15), (15, 16), (17, 18), (18, 19), (19, 20)]

mp_hands = mp.solutions.hands
hands = mp_hands.Hands(max_num_hands=2,min_detection_confidence=0.7,min_tracking_confidence=0.7)

QUESTION = ["a", "b", "c", "d", "e", "f", "g", "h", "i", "j", "k", "l", "m", "n",
            "o", "p", "q", "r", "s", "t", "u", "v", "w", "x", "y", "z"]

def write_score(num,score):
    wb = xl.load_workbook("score.xlsx")
    sheet = wb.active
    if num == 0:
        high_score = sheet["C2"].value
        if score > high_score:
            sheet["A2"].value = score
    if num == 1:
        high_score = sheet["C2"].value
        if score > high_score:
            sheet["B2"].value = score
    if num == 2:
        high_score = sheet["C2"].value
        if high_score == 0:
            sheet["C2"].value = score
        elif score < high_score:
            sheet["C2"].value = score
    if num == 3:
        high_score = sheet["C2"].value
        if score > high_score:
            sheet["D2"].value = score
    wb.save("score.xlsx")
    wb.close()

class GUI_Main(tk.Frame):
    def __init__(self,master):
        super().__init__(master)
        #self.pack()

        self.master.geometry("700x700")
        self.master.title("メイン画面")

        self.create_widgets()

    def create_widgets(self):
        self.canvas1 = tk.Canvas(self.master, bg = "skyblue", height = 700, width = 700)
        self.canvas1.place(x=0,y=0)

        self.EndButton = tk.Button(self.master, text = 'ゲームおしまい', width = 0, height = 0, command = self.end_button)
        self.EndButton.place(x=40)

        self.ScoreBtton = tk.Button(self.master, text = 'スコア画面', width = 0, height = 0, command = self.Score_button)
        self.ScoreBtton.place(x=600)

        self.GoJanken = tk.Button(self.master, text = 'じゃんけん', width = 90, height = 5, command = self.Go_Janken)
        self.GoJanken.place(x=30,y=100)

        self.GoAtti = tk.Button(self.master, text = 'あっちむいてほい', width = 90, height = 5, command = self.Go_Atti)
        self.GoAtti.place(x=30,y=250)

        self.GoTyping = tk.Button(self.master, text = '1文字タイピング', width = 90, height = 5, command = self.Go_Typing)
        self.GoTyping.place(x=30,y=400)

        self.GoFace = tk.Button(self.master, text = '顔文字ぴったん', width = 90, height = 5, command = self.Go_Face)
        self.GoFace.place(x=30,y=550)

    def end_button(self):
        sys.exit()

    def Score_button(self):
        GUI_Score(master = self.master)

    def Go_Janken(self):
        GUI_Janken(master = self.master)
        self.destroy()

    def Go_Atti(self):
        GUI_Atti(master = self.master)
        self.destroy()

    def Go_Typing(self):
        GUI_Typing(master = self.master)

    def Go_Face(self):
        GUI_Face(master = self.master)


class GUI_Score(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        self.pack()

        self.master.geometry("700x700")
        self.master.title("スコア画面")
        self.score_load()
        self.create_widgets()

    def create_widgets(self):
        self.canvas1 = tk.Canvas(self.master, bg = "#F0F0F0", height = 700, width = 700)
        self.canvas1.place(x=0,y=0)
        self.janken_score_text = tk.StringVar()
        self.janken_score_text.set(Janken_high_score)
        self.janken_score_label = tk.Label(self.master,text = Janken_high_score)
        self.janken_score_label.place(x=400,y=200)
        self.janken_score_label = tk.Label(self.master,text = "じゃんけんのハイスコアは")
        self.janken_score_label.place(x=200,y=200)

        self.atti_score_text = tk.StringVar()
        self.atti_score_text.set(Atti_high_score)
        self.atti_score_label = tk.Label(self.master,text = Atti_high_score)
        self.atti_score_label.place(x=400,y=250)
        self.janken_score_label = tk.Label(self.master,text = "あっちむいてほいのハイスコアは")
        self.janken_score_label.place(x=200,y=250)

        self.typing_score_text = tk.StringVar()
        self.typing_score_text.set(Typing_high_score)
        self.typing_score_label = tk.Label(self.master,text = Typing_high_score)
        self.typing_score_label.place(x=400,y=300)
        self.janken_score_label = tk.Label(self.master,text = "1文字タイピングのハイスコアは")
        self.janken_score_label.place(x=200,y=300)

        self.face_score_text = tk.StringVar()
        self.face_score_text.set(Face_high_score)
        self.face_score_label = tk.Label(self.master,text = Face_high_score)
        self.face_score_label.place(x=400,y=350)
        self.janken_score_label = tk.Label(self.master,text = "顔文字ぴったんのハイスコアは")
        self.janken_score_label.place(x=200,y=350)

        self.Back_Main = tk.Button(self.master, text = 'メイン画面に戻る', width = 0, height = 0, command = self.back_main_button)
        self.Back_Main.place(x=40)
        self.Back_Main = tk.Button(self.master, text = 'スコアリセット', width = 0, height = 0, command = self.score_reset)
        self.Back_Main.place(x=600)

    def score_load(self):
        global Janken_high_score,Atti_high_score,Typing_high_score,Face_high_score
        wb = xl.load_workbook("score.xlsx")
        sheet = wb.active
        Janken_high_score = int(sheet["A2"].value)
        Atti_high_score = int(sheet["B2"].value)
        Typing_high_score = int(sheet["C2"].value)
        Face_high_score = int(sheet["D2"].value)
        wb.close()

    def back_main_button(self):
        #GUI_Score.destroy(self)#これで消えそう(消えなかった)
        GUI_Main(master = self.master)

    def score_reset(self):
        wb = xl.load_workbook("score.xlsx")
        sheet = wb.active
        sheet["A2"].value = 0
        sheet["B2"].value = 0
        sheet["C2"].value = 0
        sheet["D2"].value = 0
        wb.save("score.xlsx")
        wb.close()

class GUI_Janken(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        self.pack()

        self.master.geometry("700x700")
        self.master.title("じゃんけん")

        global JANKEN_EVENT_TIMES,index,janken_game_count,win_count,confirm_count,start_count,img,img_gu,img_pa,img_choki,aiko_flag
        JANKEN_EVENT_TIMES = 10 #じゃんけん回数をいじれる

        self.img_gu = tk.PhotoImage(file = "dataset/gu.png", width = 500, height = 500)
        self.img_choki = tk.PhotoImage(file = "dataset/choki.png", width = 500, height = 500)
        self.img_pa = tk.PhotoImage(file = "dataset/pa.png", width = 500, height = 500)

        self.create_widgets()

        self.cap = cv2.VideoCapture(0)
        #self.capread()#処理を速くするために事前にreadしておく

    def end():
        root.destroy()

    def back_main_button(self,event):
        self.cap.release()
        GUI_Main(master = self.master)

    def capread(self):
        global success,media_img,results
        self.success, self.media_img = self.cap.read()
        self.media_img = cv2.flip(self.media_img, 1)
        self.results = hands.process(cv2.cvtColor(self.media_img, cv2.COLOR_BGR2RGB))

    def create_widgets(self):
        global canvas1
        self.canvas1 = tk.Canvas(self.master, bg = "#F0F0F0", height = 700, width = 700)
        self.canvas1.place(x=0,y=0)

        self.canvas1.create_image(30,30,image='', anchor = tk.NW,tag ='im')

        self.Back_Main = tk.Button(self.master, text = 'メイン画面に戻る', width = 0, height = 0)
        self.Back_Main.place(x=40)
        self.Back_Main.bind("<Button-1>",self.back_main_button)
        self.GameStart_Button = tk.Button(self.master, text = 'ゲームスタートボタン', width = 0, height = 0)
        self.GameStart_Button.place(x=200)
        self.GameStart_Button.bind("<Button-1>",self.start_timer)

        self.text1 = tk.StringVar()
        self.text1.set("あなたが出したのは")
        self.label1 = tk.Label(self.master,textvariable = self.text1)
        self.label1.place(x=300,y=300)
        self.text2 = tk.StringVar()
        self.text2.set("")
        self.label2 = tk.Label(self.master,textvariable = self.text2)
        self.label2.place(x=300,y=320)
        self.text3 = tk.StringVar()
        self.text3.set("")
        self.label3 = tk.Label(self.master,textvariable = self.text3)
        self.label3.place(x=300,y=340)
        self.text4 = tk.StringVar()
        self.text4.set("")
        self.label4 = tk.Label(self.master,textvariable = self.text4)
        self.label4.place(x=300,y=360)
        self.text5 = tk.StringVar()
        self.text5.set("")
        self.label5 = tk.Label(self.master,textvariable = self.text5)
        self.label5.place(x=300,y=360)


    def start_state(self):
        self.confirm_count = 0
        self.start_count = 0
        self.win_count = 0
        self.janken_game_count = 1
        self.aiko_flag = 0
        self.text2.set("")
        self.text3.set("")
        self.text4.set("")
        self.text5.set("")

    def start_timer(self,event):
        self.start_state()#初期化
        self.text1.set("勝った回数は："+ str(self.win_count)+"　/　勝負回数は：" + str(self.janken_game_count))
        self.playsound()
        self.after(1700,self.random_img)
        self.after(1725,self.janken)

    def start_timer2(self):
        self.confirm_count = 0
        self.start_count = 0
        self.text1.set("勝った回数は："+ str(self.win_count)+"　/　勝負回数は：" + str(self.janken_game_count))
        self.playsound()
        self.after(1700,self.random_img)
        self.after(1725,self.janken)

    def aiko_state(self):
        self.after(1700,self.random_img)
        self.after(1725,self.janken)

    def reStart(self):
        if self.start_count ==  0:
            self.start_count = 1
            self.janken_game_count += 1
        if self.janken_game_count <= JANKEN_EVENT_TIMES:
            self.after(1000,self.start_timer2)
        else:
            self.after(1,self.display_result)

    def playsound(self):
        if self.aiko_flag == 0:
            sound = lambda: winsound.PlaySound("dataset/jankenpon_01.wav", winsound.SND_FILENAME)
            thread_playsound_jankenpon = threading.Thread(target = sound)
            thread_playsound_jankenpon.start()
        else:
            sound = lambda: winsound.PlaySound("dataset/aikodesho_01.wav", winsound.SND_FILENAME)
            thread_playsound_aiko = threading.Thread(target = sound)
            thread_playsound_aiko.start()
            self.aiko_flag = 0

    def win_state(self):
        self.text3.set("勝ち")
        if self.confirm_count == 0:
             self.confirm_count = 1
             self.win_count += 1

    def display_result(self):
        write_score(0,self.win_count)
        self.text5.set("おしまい")


    def random_img(self):
        self.index = random.randint(0,2)
        self.canvas1.delete('im')#一旦削除してimgを上書き
        if self.index == 0:
            self.canvas1.create_image(30,30,image=self.img_gu, anchor = tk.NW,tag = 'im')
        elif self.index == 1:
            self.canvas1.create_image(30,30,image=self.img_choki, anchor = tk.NW,tag = 'im')
        else:
            self.canvas1.create_image(30,30,image=self.img_pa, anchor = tk.NW,tag = 'im')

    def janken(self):#mediapipe
        self.success, self.media_img = self.cap.read()
        self.cap.set(cv2.CAP_PROP_FOURCC, cv2.VideoWriter_fourcc('H', '2', '6', '4')); #高速化（あまり意味ない？
        self.media_img = cv2.flip(self.media_img, 1)
        self.results = hands.process(cv2.cvtColor(self.media_img, cv2.COLOR_BGR2RGB))
        self.text2.set("読み取れませんでした。もう一度やり直してください")
        if self.results.multi_hand_landmarks:
            # 検出した手の数分繰り返し
            for self.h_id, self.hand_landmarks in enumerate(self.results.multi_hand_landmarks):
                def calcDistance(p0,p1):
                    a1 = p1.x - p0.x
                    a2 = p1.y - p0.y
                    return math.sqrt(a1*a1+a2*a2)
                self.index_finger_dis = calcDistance(self.hand_landmarks.landmark[0],self.hand_landmarks.landmark[8])
                self.pinky_finger_dis = calcDistance(self.hand_landmarks.landmark[0],self.hand_landmarks.landmark[20])
                if self.index_finger_dis < 0.35 and self.pinky_finger_dis < 0.3:
                    self.text2.set("あなたはグーを出した")
                    hand_value = 0
                    break
                elif self.index_finger_dis > 0.35 and self.pinky_finger_dis < 0.3:
                    self.text2.set("あなたはチョキを出した")
                    hand_value = 1
                    break
                elif self.index_finger_dis > 0.35 and self.pinky_finger_dis > 0.3:
                    self.text2.set("あなたはパーを出した")
                    hand_value = 2
                    break
            if hand_value == self.index:
                self.text3.set("あいこ")
                self.aiko_flag = 1
                self.playsound()
                self.aiko_state()
            elif hand_value == 0 and self.index == 1:
                self.win_state()
                self.reStart()
            elif hand_value == 0 and self.index == 2:
                self.text3.set("負け")
                self.reStart()
            elif hand_value == 1 and self.index == 0:
                self.text3.set("負け")
                self.reStart()
            elif hand_value == 1 and self.index == 2:
                self.win_state()
                self.reStart()
            elif hand_value == 2 and self.index == 0:
                self.win_state()
                self.reStart()
            else:
                self.text3.set("負け")
                self.reStart()

class GUI_Atti(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        self.pack()

        self.master.geometry("700x700")
        self.master.title("あっちむいてほい")

        global ATTI_EVENT_TIMES,index,atti_game_count,win_count,confirm_count,start_count,img,img_front,img_up,img_down,img_left,img_right
        ATTI_EVENT_TIMES = 10
        self.img_front = tk.PhotoImage(file = "dataset/front.png", width = 500, height = 500)
        self.img_up = tk.PhotoImage(file = "dataset/up.png", width = 500, height = 500)
        self.img_down = tk.PhotoImage(file = "dataset/down.png", width = 500, height = 500)
        self.img_left = tk.PhotoImage(file = "dataset/left.png", width = 500, height = 500)
        self.img_right = tk.PhotoImage(file = "dataset/right.png", width = 500, height = 500)

        self.create_widgets()

        self.cap = cv2.VideoCapture(0)
        self.capread()#処理を軽くするために事前にreadしておく


    def end():
        root.destroy()

    def back_main_button(self,event):
        self.cap.release()
        GUI_Main(master = self.master)

    def capread(self):
        global success,media_img,results
        self.success, self.media_img = self.cap.read()
        self.media_img = cv2.flip(self.media_img, 1)
        self.results = hands.process(cv2.cvtColor(self.media_img, cv2.COLOR_BGR2RGB))

    def create_widgets(self):
        global canvas1
        self.canvas1 = tk.Canvas(self.master, bg = "#F0F0F0", height = 700, width = 700)
        self.canvas1.place(x=0,y=0)

        self.canvas1.create_image(30,30,image=self.img_front, anchor = tk.NW,tag ='im')

        self.Back_Main = tk.Button(self.master, text = 'メイン画面に戻る', width = 0, height = 0)
        self.Back_Main.place(x=40)
        self.Back_Main.bind("<Button-1>",self.back_main_button)
        self.GameStart_Button = tk.Button(self.master, text = 'ゲームスタートボタン', width = 0, height = 0)
        self.GameStart_Button.place(x=200)
        self.GameStart_Button.bind("<Button-1>",self.start_timer)

        self.text1 = tk.StringVar()
        self.text1.set("あなたが出したのは")
        self.label1 = tk.Label(self.master,textvariable = self.text1)
        self.label1.place(x=300,y=300)
        self.text2 = tk.StringVar()
        self.text2.set("")
        self.label2 = tk.Label(self.master,textvariable = self.text2)
        self.label2.place(x=300,y=320)
        self.text3 = tk.StringVar()
        self.text3.set("")
        self.label3 = tk.Label(self.master,textvariable = self.text3)
        self.label3.place(x=300,y=340)
        self.text4 = tk.StringVar()
        self.text4.set("")
        self.label4 = tk.Label(self.master,textvariable = self.text4)
        self.label4.place(x=300,y=360)
        self.text5 = tk.StringVar()
        self.text5.set("")
        self.label5 = tk.Label(self.master,textvariable = self.text5)
        self.label5.place(x=300,y=360)

    def start_state(self):
        self.confirm_count = 0
        self.start_count = 0
        self.win_count = 0
        self.atti_game_count = 1
        self.text2.set("")
        self.text3.set("")
        self.text4.set("")
        self.text5.set("")


    def start_timer(self,event):
        self.start_state()#初期化
        self.text1.set("指が合った回数は："+ str(self.win_count)+"　/　勝負回数は：" + str(self.atti_game_count))
        self.playsound()
        self.after(1400,self.random_img)
        self.after(1425,self.atti)

    def start_timer2(self):
        self.confirm_count = 0
        self.start_count = 0
        self.text1.set("指が合った回数は："+ str(self.win_count)+"　/　勝負回数は：" + str(self.atti_game_count))
        self.playsound()
        self.after(1400,self.random_img)
        self.after(1425,self.atti)


    def reStart(self):
        if self.start_count ==  0:
            self.start_count = 1
            self.atti_game_count += 1
        if self.atti_game_count <= ATTI_EVENT_TIMES:
            self.after(1000,self.start_timer2)
        else:
            self.after(1,self.display_result)

    def playsound(self):
        sound = lambda: winsound.PlaySound("dataset/atti.wav", winsound.SND_FILENAME)
        thread_playsound_atti = threading.Thread(target = sound)
        thread_playsound_atti.start()

    def match_state(self):
        self.text3.set("一致")
        if self.confirm_count == 0:
             self.confirm_count = 1
             self.win_count += 1

    def display_result(self):
        write_score(1,self.win_count)
        self.text5.set("おしまい")

    def random_img(self):
        self.index = random.randint(0,3)
        self.canvas1.delete('im')#一旦削除してimgを上書き
        if self.index == 0:
            self.canvas1.create_image(30,30,image=self.img_up, anchor = tk.NW,tag = 'im')
        elif self.index == 1:
            self.canvas1.create_image(30,30,image=self.img_down, anchor = tk.NW,tag = 'im')
        elif self.index == 2:
            self.canvas1.create_image(30,30,image=self.img_left, anchor = tk.NW,tag = 'im')
        else:
            self.canvas1.create_image(30,30,image=self.img_right, anchor = tk.NW,tag = 'im')

    def atti(self):#mediapipe
        self.success, self.media_img = self.cap.read()
        self.cap.set(cv2.CAP_PROP_FOURCC, cv2.VideoWriter_fourcc('H', '2', '6', '4')); #高速化（あまり意味ない？
        self.media_img = cv2.flip(self.media_img, 1)
        self.results = hands.process(cv2.cvtColor(self.media_img, cv2.COLOR_BGR2RGB))
        self.text2.set("読み取れませんでした。")
        if self.results.multi_hand_landmarks:
            # 検出した手の数分繰り返し
            for self.h_id, self.hand_landmarks in enumerate(self.results.multi_hand_landmarks):
                self.x_index_finger = self.hand_landmarks.landmark[8].x
                self.y_index_finger = self.hand_landmarks.landmark[8].y
                print("x："+str(self.x_index_finger) + "  ,  "+"y：" + str(self.y_index_finger) )
                if(0.7 > self.y_index_finger > 0.4 and 0.6 > self.x_index_finger > 0.4):
                    self.hand_value = 0
                    self.text2.set("上です")
                    print("up")
                elif(self.y_index_finger > 0.7 and 0.6 > self.x_index_finger > 0.4):
                    self.hand_value = 1
                    self.text2.set("下です")
                    print("down")
                elif(0.9 > self.y_index_finger > 0.5 and 0.5 > self.x_index_finger):
                    self.hand_value = 2
                    self.text2.set("左です")
                    print("left")
                elif(0.9 > self.y_index_finger > 0.6 and 0.9 > self.x_index_finger > 0.6):#判定がしびあ
                    self.hand_value = 3
                    self.text2.set("右です")
                    print("right")

            if self.hand_value == self.index:
                self.match_state()
                self.reStart()
            else:
                self.text3.set("不一致")
                self.reStart()

class GUI_Typing(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        self.pack()

        master.geometry("700x700")
        master.title("一文字タイピング")

        #問題数インデックス
        self.index = 0
        self.index2 = random.randint(0,25)

        #正解数カウント用
        self.correct_cnt = 0
        self.start_button_flag = False

        self.create_widgets()

        #経過時間スレッドの開始


        #Tkインスタンスに対してキーイベント処理を実装
        self.master.bind("<KeyPress>", self.type_event)

    #ウィジェットの生成と配置
    def create_widgets(self):
        self.canvas1 = tk.Canvas(self.master, bg = "#F0F0F0", height = 700, width = 700)
        self.canvas1.place(x=0,y=0)
        self.q_label = tk.Label(self.master, text="問題：", font=("",20))
        self.q_label.place(x=40,y=100)
        self.q_label2 = tk.Label(self.master, text=QUESTION[self.index2], width=10, anchor="w", font=("",20))
        self.q_label2.place(x=40,y=150)
        self.ans_label = tk.Label(self.master, text="解答：", font=("",20))
        self.ans_label.place(x=40,y=200)
        self.ans_label2 = tk.Label(self.master, text="", width=10, anchor="w", font=("",20))
        self.ans_label2.place(x=40,y=250)
        self.result_label = tk.Label(self.master, text="", font=("",20))
        self.result_label.place(x=40,y=300)

        #時間計測用のラベル
        self.time_label = tk.Label(self.master, text="", font=("",20))
        self.time_label.place(x=40,y=350)

        self.Back_Main = tk.Button(self.master.master, text = 'メイン画面に戻る', width = 0, height = 0)
        self.Back_Main.place(x=40)
        self.Back_Main.bind("<Button-1>",self.back_main_button)

        self.GameStart_Button = tk.Button(self.master, text = 'ゲームスタートボタン', width = 0, height = 0, command = self.start_timer)
        self.GameStart_Button.place(x=200)


        self.flg2 = True

    def start_timer(self):
        self.start_button_flag = True
        t = threading.Thread(target=self.timer)
        t.start()

    def back_main_button(self,event):
        GUI_Main(master = self.master)

    #キー入力時のイベント処理
    def type_event(self, event):
        if self.start_button_flag == True:
            if self.flg == True:
                #入力値がEnterの場合は答え合わせ
                if event.keysym == "Return":
                    if self.q_label2["text"] == self.ans_label2["text"]:
                        sound = lambda: winsound.PlaySound("dataset/pinpon.wav", winsound.SND_FILENAME)
                        thread_playsound_atti = threading.Thread(target = sound)
                        thread_playsound_atti.start()
                        self.result_label.configure(text="ピンポン！", fg="red")
                        self.correct_cnt += 1
                        i = 1
                    else:
                        sound = lambda: winsound.PlaySound("dataset/bubuu.wav", winsound.SND_FILENAME)
                        thread_playsound_atti = threading.Thread(target = sound)
                        thread_playsound_atti.start()
                        self.result_label.configure(text="ブブー！", fg="blue")


                    #解答欄をクリア
                    self.ans_label2.configure(text="")

                    #次の問題を出題
                    self.index += 1
                    if self.correct_cnt == 10:
                        self.flg = False
                        sound = lambda: winsound.PlaySound("dataset/clear.wav", winsound.SND_FILENAME)
                        thread_playsound_atti = threading.Thread(target = sound)
                        thread_playsound_atti.start()
                        second_score = self.second
                        #print(second_score)
                        write_score(2,second_score)
                        #self.q_label2.configure(text="終了！")
                        #messagebox.showinfo("リザルト", f"あなたのスコアは{self.correct_cnt}/{self.index}問正解です。")
                        #sys.exit(0)

                    self.index2 = random.randint(0,14)
                    self.q_label2.configure(text=QUESTION[self.index2])

                elif event.keysym == "BackSpace":
                    text = self.ans_label2["text"]
                    self.ans_label2["text"] = text[:-1]
                else:
                    #入力値がEnter以外の場合は文字入力としてラベルに追記する
                    self.ans_label2["text"] += event.keysym

    def timer(self):
        global flg
        if self.start_button_flag == True:
            self.second = 0
            self.flg = True
            while self.flg:
                self.second += 1
                self.time_label.configure(text=f"経過時間：{self.second}秒")
                time.sleep(1)

class GUI_Face(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        self.pack()

        self.master.geometry("700x700")
        self.master.title("精度の悪い、顔ぴったん")

        global FACE_EVENT_TIMES,index,face_game_count,match_count,confirm_count,start_count,face_value,face_cascade,face_parts_detector,close_eye,open_mouth,up_mouth,count
        FACE_EVENT_TIMES = 5
        self.img_normal = tk.PhotoImage(file = "dataset/normal.png", width = 500, height = 500)
        self.img_happy = tk.PhotoImage(file = "dataset/happy.png", width = 500, height = 500)
        self.img_angry = tk.PhotoImage(file = "dataset/angry.png", width = 500, height = 500)
        self.img_cry = tk.PhotoImage(file = "dataset/cry.png", width = 500, height = 500)
        self.img_surprise =tk.PhotoImage(file = "dataset/surprise.png", width = 500, height = 500)

        self.close_eye = 0.0
        self.open_mouth = 0.0
        self.up_mouth = 0.0
        self.count = 0

        self.face_cascade = cv2.CascadeClassifier('dataset/haarcascade_frontalface_alt2.xml')
        self.face_parts_detector = dlib.shape_predictor('dataset/shape_predictor_68_face_landmarks.dat')

        self.create_widgets()

        self.cap = cv2.VideoCapture(0)
        self.capread()#処理を軽くするために事前にreadしておく

    def end():
        root.destroy()

    def back_main_button(self,event):
        self.cap.release()
        GUI_Main(master = self.master)

    def capread(self):
        global success,media_img,results
        self.ret, self.rgb = self.cap.read()
        self.gray = cv2.cvtColor(self.rgb, cv2.COLOR_RGB2GRAY)
        self.faces = self.face_cascade.detectMultiScale(
            self.gray, scaleFactor=1.11, minNeighbors=5, minSize=(100,100))

    def create_widgets(self):
        global canvas1
        self.canvas1 = tk.Canvas(self.master, bg = "#F0F0F0", height = 700, width = 700)
        self.canvas1.place(x=0,y=0)

        self.canvas1.create_image(30,30,image='', anchor = tk.NW,tag ='im')

        self.Back_Main = tk.Button(self.master, text = '終わる', width = 0, height = 0)
        self.Back_Main.place(x=40)
        self.Back_Main.bind("<Button-1>",self.back_main_button)
        self.GameStart_Button = tk.Button(self.master, text = 'ゲームスタートボタン', width = 0, height = 0)
        self.GameStart_Button.place(x=200)
        self.GameStart_Button.bind("<Button-1>",self.start_timer)

        self.text1 = tk.StringVar()
        self.text1.set("あなたの顔は")
        self.label1 = tk.Label(self.master,textvariable = self.text1)
        self.label1.place(x=300,y=300)
        self.text2 = tk.StringVar()
        self.text2.set("")
        self.label2 = tk.Label(self.master,textvariable = self.text2)
        self.label2.place(x=300,y=320)
        self.text3 = tk.StringVar()
        self.text3.set("")
        self.label3 = tk.Label(self.master,textvariable = self.text3)
        self.label3.place(x=300,y=340)
        self.text4 = tk.StringVar()
        self.text4.set("")
        self.label4 = tk.Label(self.master,textvariable = self.text4)
        self.label4.place(x=300,y=360)
        self.text5 = tk.StringVar()
        self.text5.set("")
        self.label5 = tk.Label(self.master,textvariable = self.text5)
        self.label5.place(x=300,y=360)

    def start_state(self):
        self.confirm_count = 0
        self.start_count = 0
        self.match_count = 0
        self.face_game_count = 1
        self.text2.set("")
        self.text3.set("")
        self.text4.set("")
        self.text5.set("")


    def start_timer(self,event):
        self.start_state()#初期化
        self.text1.set("合った回数は："+ str(self.match_count)+"　/　勝負回数は：" + str(self.face_game_count))
        self.playsound()
        self.after(500,self.random_img)
        self.after(1725,self.match_face)

    def start_timer2(self):
        self.confirm_count = 0
        self.start_count = 0
        self.text1.set("合った回数は："+ str(self.match_count)+"　/　勝負回数は：" + str(self.face_game_count))
        self.playsound()
        self.after(1400,self.random_img)
        self.after(1425,self.match_face)

    def playsound(self):
        sound = lambda: winsound.PlaySound("dataset/atti.wav", winsound.SND_FILENAME)
        thread_playsound_atti = threading.Thread(target = sound)
        thread_playsound_atti.start()

    def reStart(self):
        if self.start_count ==  0:
            self.start_count = 1
            self.face_game_count += 1
        if self.face_game_count <= FACE_EVENT_TIMES:
            self.after(1000,self.start_timer2)
        else:
            self.after(1,self.display_result)

    def match_state(self):
        self.text3.set("一致")
        if self.confirm_count == 0:
             self.confirm_count = 1
             self.match_count += 1

    def display_result(self):
        write_score(3,self.match_count)
        self.text5.set("おしまい")

    def random_img(self):
        self.index = random.randint(0,3)
        self.canvas1.delete('im')#一旦削除してimgを上書き
        if self.index == 0:
            self.canvas1.create_image(30,30,image=self.img_happy, anchor = tk.NW,tag = 'im')
        elif self.index == 1:
            self.canvas1.create_image(30,30,image=self.img_angry, anchor = tk.NW,tag = 'im')
        elif self.index == 2:
            self.canvas1.create_image(30,30,image=self.img_cry, anchor = tk.NW,tag = 'im')
        else:
            self.canvas1.create_image(30,30,image=self.img_surprise, anchor = tk.NW,tag = 'im')

    def match_face(self):#mediapipe
        global index
        self.ret, self.rgb = self.cap.read()
        self.gray = cv2.cvtColor(self.rgb, cv2.COLOR_RGB2GRAY)
        self.faces = self.face_cascade.detectMultiScale(
            self.gray, scaleFactor=1.11, minNeighbors=5, minSize=(100,100))
        self.text2.set("読み取れませんでした。")
        start_time = 0
        end_time = 0
        if len(self.faces) == 1:
            start_time = time.time()
            x, y, w, h = self.faces[0, :]
            while True:
                end_time = time.time()
                self.face = dlib.rectangle(x, y, x + w, y + h)
                self.face_parts = self.face_parts_detector(self.gray, self.face)
                self.face_parts = face_utils.shape_to_np(self.face_parts)
                def calc_smile(mouth_value):
                    A = distance.euclidean(mouth_value[0],mouth_value[6])
                    B = distance.euclidean(mouth_value[3],mouth_value[9])
                    smile = (A+B)/2
                    return round(smile,3)
                def mouth2_value(mouth):
                    A = distance.euclidean(mouth[14],mouth[18])
                    return round(A,3)
                def calc_ear(eye):
                    A = distance.euclidean(eye[1], eye[5])
                    B = distance.euclidean(eye[2], eye[4])
                    C = distance.euclidean(eye[0], eye[3])
                    eye_ear = (A + B) / (2.0 * C)
                    return round(eye_ear, 3)
                smile = calc_smile(self.face_parts[48:68])
                mouth2 = mouth2_value(self.face_parts[48:68])
                left_eye_ear = calc_ear(self.face_parts[42:48])#左目の振り分け番号
                right_eye_ear = calc_ear(self.face_parts[36:42])#右眼の振り分け番号

                a = distance.euclidean(self.face_parts[37],self.face_parts[49])
                b = distance.euclidean(self.face_parts[46],self.face_parts[55])
                c = (a*b)/(a+b)#笑顔のもとになる値
                if mouth2 > 10 and (left_eye_ear + right_eye_ear) < 0.35:#あくび判定
                    self.face_value = 2
                    self.text2.set("cry")
                    break
                elif (left_eye_ear+right_eye_ear) < 0.35:
                    self.face_value = 1
                    self.text2.set("angry")
                    break
                elif mouth2 > 10 and (left_eye_ear + right_eye_ear) > 0.35:#あくび判定
                    self.face_value = 3
                    self.text2.set("surprise")
                    break
                elif smile > c and mouth2 < 15:#スマイル判定
                    self.face_value = 0
                    self.text2.set("smile")
                    break
                elif (end_time-start_time > 1.0):#１秒後にbreak
                    break
            if self.face_value == self.index:
                self.match_state()
                self.reStart()
            else:
                self.text3.set("不一致")
                self.reStart()


def main():
    root = tk.Tk()
    app = GUI_Main(master=root)#Inherit

    app.mainloop()

if __name__ == "__main__":
    main()
